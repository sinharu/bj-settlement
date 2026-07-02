import re
import zipfile
from pathlib import Path
from io import BytesIO

import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

from processor import process_dataframe


st.set_page_config(page_title="BJ 하트 집계", layout="centered")


# ==================================================
# 🔐 비밀번호 게이트
# ==================================================
def check_password():
    def password_entered():
        if st.session_state.get("password", "") == st.secrets["APP_PASSWORD"]:
            st.session_state["password_correct"] = True
            st.session_state.pop("password", None)
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        st.text_input("비밀번호를 입력하세요", type="password", key="password", on_change=password_entered)
        return False

    if not st.session_state["password_correct"]:
        st.text_input("비밀번호를 입력하세요", type="password", key="password", on_change=password_entered)
        st.error("비밀번호가 틀렸습니다.")
        return False

    return True


if not check_password():
    st.stop()


# ==================================================
# 📦 엑셀 공통 유틸 (콤마/테두리/열너비)
# ==================================================
thin = Side(style="thin")
all_border = Border(left=thin, right=thin, top=thin, bottom=thin)

def apply_border(ws):
    # "값이 있는 셀" 전부 테두리
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value not in (None, ""):
                cell.border = all_border

def auto_width(ws, min_w=18, max_w=45, pad=4):
    # 기본 넓이 유지 + 데이터 길이 따라 자동 확장
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value not in (None, ""):
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + pad, min_w), max_w)

def format_header_row(ws, header_row=1):
    # 헤더 정렬(선택)
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=c)
        if cell.value not in (None, ""):
            cell.alignment = Alignment(horizontal="center")


# ==================================================
# 📌 화면 시작
# ==================================================
st.title("BJ 하트 집계 (BJ 전달용)")
st.caption("CSV / XLSX 업로드 → 웹 요약표 확인 → BJ별 엑셀 다운로드")

uploaded_files = st.file_uploader(
    "CSV 또는 XLSX 파일을 업로드하세요",
    type=["csv", "xlsx"],
    accept_multiple_files=True
)

if not uploaded_files:
    st.info("파일을 업로드하면 집계 결과가 표시됩니다.")
    st.stop()


# ==================================================
# 📥 파일 읽기
# ==================================================
dfs = []
MAX_STANDARD_ROUNDS = 15
standard_round_count = min(len(uploaded_files), MAX_STANDARD_ROUNDS)
round_labels = [f"{idx}회차" for idx in range(1, standard_round_count + 1)]
for idx, f in enumerate(uploaded_files, start=1):
    try:
        if f.name.lower().endswith(".csv"):
            df = pd.read_csv(f)
        else:
            df = pd.read_excel(f)
        if len(uploaded_files) > 1:
            if len(uploaded_files) > MAX_STANDARD_ROUNDS:
                round_no = ((idx - 1) // 2) + 1
            else:
                round_no = idx
            round_no = min(round_no, MAX_STANDARD_ROUNDS)
            df["업로드회차"] = f"{round_no}회차"
        dfs.append(df)
    except Exception as e:
        st.error(f"{f.name} 읽기 실패: {e}")

if not dfs:
    st.error("읽을 수 있는 파일이 없습니다.")
    st.stop()

merged = pd.concat(dfs, ignore_index=True)


# ==================================================
# 📅 파일 1개 업로드 시 날짜 prefix (파일명 우선 → 없으면 데이터 최솟날짜)
# ==================================================
def extract_prefix_from_filename(files):
    for f in files:
        stem = Path(f.name).stem
        m = re.match(r"^(\d{2}\.\d{2})", stem)
        if m:
            return m.group(1)
    return None

def extract_earliest_date_prefix(df):
    col_time = next((c for c in df.columns if "후원" in c and "시간" in c), None)
    if not col_time:
        return None
    tmp = df[[col_time]].copy()
    tmp[col_time] = pd.to_datetime(tmp[col_time], errors="coerce")
    min_dt = tmp[col_time].min()
    if pd.isna(min_dt):
        return None
    return min_dt.strftime("%m.%d")

if len(uploaded_files) == 1:
    prefix = extract_prefix_from_filename(uploaded_files)
    if not prefix:
        prefix = extract_earliest_date_prefix(merged)
else:
    prefix = None  # 여러개면 prefix 안 붙임


# ==================================================
# 🧠 하트 구분 / 아이디-닉네임 분리
# ==================================================
def classify_heart_type(user_id: str) -> str:
    s = str(user_id)
    if "@ka" in s:
        return "일반"
    if "@" in s:
        return "제휴"
    return "일반"

def split_id_nickname(text):
    text = str(text)
    if "(" in text and ")" in text:
        id_part, nick_part = text.split("(", 1)
        nick_part = nick_part.rstrip(")")
    else:
        id_part = text
        nick_part = ""
    return id_part.strip(), nick_part.strip()


# ==================================================
# 📊 웹 요약표 (참여BJ별 일반/제휴/총합)
# ==================================================
try:
    tmp = merged.copy()

    col_id = next((c for c in tmp.columns if "후원" in c and "아이디" in c), None)
    col_heart = next((c for c in tmp.columns if "후원" in c and "하트" in c), None)
    col_bj = next((c for c in tmp.columns if "참여" in c and "BJ" in c), None)

    if not (col_id and col_heart and col_bj):
        st.warning("요약표: 필수 컬럼(후원아이디/후원하트/참여BJ)을 찾지 못했습니다.")
    else:
        tmp[col_heart] = pd.to_numeric(tmp[col_heart], errors="coerce").fillna(0)
        tmp.loc[tmp[col_heart] < 0, col_heart] = 0

        tmp["후원아이디"] = tmp[col_id].astype(str).str.replace(r"\(.*\)", "", regex=True).str.strip()
        tmp["구분"] = tmp["후원아이디"].apply(classify_heart_type)

        pivot = (
            tmp.groupby([col_bj, "구분"])[col_heart]
            .sum()
            .unstack(fill_value=0)
            .reset_index()
        )

        if "일반" not in pivot.columns:
            pivot["일반"] = 0
        if "제휴" not in pivot.columns:
            pivot["제휴"] = 0

        pivot["총합"] = pivot["일반"] + pivot["제휴"]
        pivot = pivot.rename(columns={col_bj: "참여BJ"})
        pivot = pivot[["참여BJ", "일반", "제휴", "총합"]].sort_values("총합", ascending=False)

        # 화면용 콤마(문자열) — 엑셀은 number_format으로 처리하니까 여기만 문자열로 OK
        for c in ["일반", "제휴", "총합"]:
            pivot[c] = pivot[c].apply(lambda x: f"{int(x):,}")

        st.subheader("요약_참여BJ_총계")
        st.dataframe(pivot.reset_index(drop=True), hide_index=True, use_container_width=True)

except Exception as e:
    st.warning(f"요약표 생성 중 오류: {e}")


# ==================================================
# 📁 BJ별 파일 생성 (정산용 / BJ용) - 콤마/테두리/열너비 적용
# ==================================================
result = process_dataframe(merged)

if not result:
    st.error("집계 결과가 없습니다.")
    st.stop()

def make_excel(df: pd.DataFrame, bj_name: str, detail_df=None) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "정산표"

    # 상단 합계
    total = int(pd.to_numeric(df["후원하트"], errors="coerce").fillna(0).sum())
    ws["A1"] = ""
    ws["B1"] = bj_name
    total_cell = ws["C1"]
    total_cell.value = total
    total_cell.number_format = "#,##0"

    # 헤더
    ws.append(["후원아이디", "닉네임", "후원하트"])
    format_header_row(ws, header_row=2)

    # 데이터
    for _, r in df.iterrows():
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=str(r.get("아이디", "")))
        ws.cell(row=row, column=2, value=str(r.get("닉네임", "")))

        heart = pd.to_numeric(r.get("후원하트", 0), errors="coerce")
        heart = 0 if pd.isna(heart) else int(heart)
        heart = max(heart, 0)

        cell = ws.cell(row=row, column=3, value=heart)
        cell.number_format = "#,##0"

    # 기본 폭(너무 좁아지는 것 방지) + 자동 보정
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    auto_width(ws, min_w=18, max_w=45, pad=4)
    apply_border(ws)

    # ==================================================
    # 📄 상세내역 시트 추가
    # ==================================================
    if detail_df is not None and not detail_df.empty:

        detail_ws = wb.create_sheet("상세내역")

        detail_ws.append([
            "날짜",
            "시간",
            "아이디",
            "닉네임",
            "하트",
            "구분"
        ])

        format_header_row(detail_ws, 1)

        detail_df = detail_df.sort_values(
            by=["날짜", "시간"],
            ascending=True
        )

        for _, r in detail_df.iterrows():

            row = detail_ws.max_row + 1

            detail_ws.cell(
                row=row,
                column=1,
                value=r.get("날짜")
            )

            detail_ws.cell(
                row=row,
                column=2,
                value=r.get("시간")
            )

            detail_ws.cell(
                row=row,
                column=3,
                value=r.get("아이디")
            )

            detail_ws.cell(
                row=row,
                column=4,
                value=r.get("닉네임")
            )

            heart_cell = detail_ws.cell(
                row=row,
                column=5,
                value=int(r.get("후원하트", 0))
            )

            heart_cell.number_format = "#,##0"

            detail_ws.cell(
                row=row,
                column=6,
                value=r.get("구분")
            )

        detail_ws.column_dimensions["A"].width = 20
        detail_ws.column_dimensions["B"].width = 18
        detail_ws.column_dimensions["C"].width = 32
        detail_ws.column_dimensions["D"].width = 26
        detail_ws.column_dimensions["E"].width = 14
        detail_ws.column_dimensions["F"].width = 12

        auto_width(detail_ws, min_w=18, max_w=45, pad=4)

        apply_border(detail_ws)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ==================================================
# 📦 총합산 파일 (여러 파일 업로드 시) - 3시트 구조
# 1) 일자별집계  2) 총합  3) BJ별 상세(각 BJ 1시트)
# ==================================================
def make_total_excel(df: pd.DataFrame) -> BytesIO | None:
    wb = Workbook()
    wb.remove(wb.active)

    tmp = df.copy()

    col_time = next((c for c in tmp.columns if "후원" in c and "시간" in c), None)
    col_idnick = next((c for c in tmp.columns if "후원" in c and "아이디" in c), None)
    col_heart = next((c for c in tmp.columns if "후원" in c and "하트" in c), None)
    col_bj = next((c for c in tmp.columns if "참여" in c and "BJ" in c), None)

    if not all([col_time, col_idnick, col_heart, col_bj]):
        return None

    tmp[col_time] = (
        tmp[col_time]
        .astype(str)
        .str.replace(r"[./]", "-", regex=True)
        .str.replace("T", " ")
        .str.strip()
    )

    tmp[col_time] = pd.to_datetime(
        tmp[col_time],
        errors="coerce",
        format="mixed"
    )

    tmp["날짜"] = tmp[col_time].dt.date
    tmp["시간"] = tmp[col_time].dt.time

    tmp[col_heart] = pd.to_numeric(tmp[col_heart], errors="coerce").fillna(0)
    tmp.loc[tmp[col_heart] < 0, col_heart] = 0
    tmp["날짜"] = tmp[col_time].dt.date
    tmp["시간"] = tmp[col_time].dt.time
    tmp[col_heart] = pd.to_numeric(tmp[col_heart], errors="coerce").fillna(0)
    tmp.loc[tmp[col_heart] < 0, col_heart] = 0

    tmp[["아이디", "닉네임"]] = tmp[col_idnick].apply(lambda x: pd.Series(split_id_nickname(x)))
    tmp["구분"] = tmp["아이디"].apply(classify_heart_type)

    # 1) 일자별집계
    ws1 = wb.create_sheet("일자별집계")
    ws1.append(["날짜", "BJ", "일반", "제휴", "총합"])
    format_header_row(ws1, 1)

    s1 = (
        tmp.groupby(["날짜", col_bj, "구분"])[col_heart]
        .sum()
        .unstack(fill_value=0)
        .reset_index()
    )
    if "일반" not in s1.columns: s1["일반"] = 0
    if "제휴" not in s1.columns: s1["제휴"] = 0
    s1["총합"] = s1["일반"] + s1["제휴"]

    for _, r in s1.iterrows():
        row = ws1.max_row + 1
        ws1.cell(row=row, column=1, value=r["날짜"])
        ws1.cell(row=row, column=2, value=r[col_bj])

        c1 = ws1.cell(row=row, column=3, value=int(r["일반"]))
        c2 = ws1.cell(row=row, column=4, value=int(r["제휴"]))
        c3 = ws1.cell(row=row, column=5, value=int(r["총합"]))
        for c in (c1, c2, c3):
            c.number_format = "#,##0"

    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 28
    ws1.column_dimensions["C"].width = 16
    ws1.column_dimensions["D"].width = 16
    ws1.column_dimensions["E"].width = 16
    auto_width(ws1, min_w=18, max_w=45, pad=4)
    apply_border(ws1)

    # 2) 총합
    ws2 = wb.create_sheet("총합")
    ws2.append(["BJ", "일반", "제휴", "총합"])
    format_header_row(ws2, 1)

    s2 = (
        tmp.groupby([col_bj, "구분"])[col_heart]
        .sum()
        .unstack(fill_value=0)
        .reset_index()
    )
    if "일반" not in s2.columns: s2["일반"] = 0
    if "제휴" not in s2.columns: s2["제휴"] = 0
    s2["총합"] = s2["일반"] + s2["제휴"]

    for _, r in s2.iterrows():
        row = ws2.max_row + 1
        ws2.cell(row=row, column=1, value=r[col_bj])

        c1 = ws2.cell(row=row, column=2, value=int(r["일반"]))
        c2 = ws2.cell(row=row, column=3, value=int(r["제휴"]))
        c3 = ws2.cell(row=row, column=4, value=int(r["총합"]))
        for c in (c1, c2, c3):
            c.number_format = "#,##0"

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 16
    auto_width(ws2, min_w=18, max_w=45, pad=4)
    apply_border(ws2)

    # 3) BJ별 상세 (각 BJ 1시트)
    for bj in tmp[col_bj].dropna().unique():
        ws = wb.create_sheet(str(bj))
        sub = tmp[tmp[col_bj] == bj].copy()

        normal_sum = int(sub[sub["구분"] == "일반"][col_heart].sum())
        partner_sum = int(sub[sub["구분"] == "제휴"][col_heart].sum())
        total_sum = normal_sum + partner_sum

        # 상단 한 줄 표시(일렬)
        ws["A1"] = "총하트"
        ws["B1"] = total_sum
        ws["C1"] = "일반하트"
        ws["D1"] = normal_sum
        ws["E1"] = "제휴하트"
        ws["F1"] = partner_sum
        ws["B1"].number_format = "#,##0"
        ws["D1"].number_format = "#,##0"
        ws["F1"].number_format = "#,##0"

        ws.append([])
        ws.append(["날짜", "시간", "아이디", "닉네임", "하트", "구분"])
        format_header_row(ws, header_row=3)

        # 정렬(원하면 여기서 날짜/시간 정렬)
        sub = sub.sort_values(by=[col_time], ascending=True)

        for _, r in sub.iterrows():
            row = ws.max_row + 1
            ws.cell(row=row, column=1, value=r["날짜"])
            ws.cell(row=row, column=2, value=r["시간"])
            ws.cell(row=row, column=3, value=r["아이디"])
            ws.cell(row=row, column=4, value=r["닉네임"])

            h = int(r[col_heart])
            hc = ws.cell(row=row, column=5, value=h)
            hc.number_format = "#,##0"

            ws.cell(row=row, column=6, value=r["구분"])

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 32
        ws.column_dimensions["D"].width = 26
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 12
        auto_width(ws, min_w=18, max_w=45, pad=4)
        apply_border(ws)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def _as_time_fraction(value) -> float:
    if pd.isna(value):
        return 1
    if hasattr(value, "hour"):
        return (value.hour * 3600 + value.minute * 60 + value.second) / 86400
    try:
        parsed = pd.to_datetime(str(value), errors="coerce")
    except Exception:
        return 1
    if pd.isna(parsed):
        return 1
    return (parsed.hour * 3600 + parsed.minute * 60 + parsed.second) / 86400


def _business_date(date_value, time_value):
    parsed = pd.to_datetime(date_value, errors="coerce")
    if pd.isna(parsed):
        return None
    business_dt = parsed
    if _as_time_fraction(time_value) < 0.625:
        business_dt = business_dt - pd.Timedelta(days=1)
    return business_dt.date()


def _save_workbook_with_cached_values(wb: Workbook, cached_values: dict[str, int | float]) -> BytesIO:
    base = BytesIO()
    wb.save(base)
    base.seek(0)

    patched = BytesIO()
    with zipfile.ZipFile(base, "r") as zin, zipfile.ZipFile(patched, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                xml = data.decode("utf-8")
                for cell_ref, value in cached_values.items():
                    if value is None:
                        continue
                    value_text = str(int(value)) if isinstance(value, float) and value.is_integer() else str(value)
                    pattern = rf'(<c\b[^>]*\br="{re.escape(cell_ref)}"[^>]*>)(.*?)(</c>)'

                    def repl(match):
                        inner = match.group(2)
                        if "<f" not in inner:
                            return match.group(0)
                        inner = re.sub(r"<v\b[^>]*/>", "", inner, flags=re.DOTALL)
                        inner = re.sub(r"<v\b[^>]*>.*?</v>", "", inner, flags=re.DOTALL)
                        inner = re.sub(r"(</f>)", rf"\1<v>{value_text}</v>", inner, count=1)
                        return match.group(1) + inner + match.group(3)

                    xml = re.sub(pattern, repl, xml, count=1, flags=re.DOTALL)
                data = xml.encode("utf-8")
            zout.writestr(item, data)

    patched.seek(0)
    return patched


def make_standard_settlement_excel(
    detail_df: pd.DataFrame,
    bj_name: str,
    all_round_labels: list[str] | None = None
) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "정산시트"
    log_ws = wb.create_sheet("후원내역")

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    header_fill = PatternFill("solid", fgColor="666666")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    input_fill = PatternFill("solid", fgColor="D9EAD3")
    title_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF")
    bold_font = Font(name="맑은 고딕", bold=True)
    normal_font = Font(name="맑은 고딕", size=11)

    def style_header(cell):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def style_input(cell):
        cell.fill = input_fill
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:G2")
    ws["A1"] = f"{bj_name} 정산표"
    ws["A1"].font = Font(name="맑은 고딕", size=14, bold=True)
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["I3"] = "정산비율"
    ws["J3"] = 0.45
    ws["I4"] = "하트단가"
    ws["J4"] = "=$J$3*100"
    ws["I5"] = "협력지원율"
    ws["J5"] = 0.05
    for cell in ("I3", "I4", "I5"):
        style_input(ws[cell])
    for cell in ("J3", "J4", "J5"):
        style_input(ws[cell])
    ws["J3"].number_format = "0%"
    ws["J4"].number_format = "#,##0"
    ws["J5"].number_format = "0%"

    sorted_detail = detail_df.copy() if detail_df is not None else pd.DataFrame()
    if not sorted_detail.empty:
        sorted_detail = sorted_detail.sort_values(by=["날짜", "시간"], ascending=True)
    if "회차" in sorted_detail.columns and sorted_detail["회차"].notna().any():
        sorted_detail["회차"] = sorted_detail["회차"].fillna("").astype(str)
        round_names = all_round_labels or sorted(
            [x for x in sorted_detail["회차"].dropna().unique() if x],
            key=lambda x: int(re.search(r"\d+", str(x)).group()) if re.search(r"\d+", str(x)) else 9999
        )
    else:
        sorted_detail["정산일자"] = sorted_detail.apply(
            lambda r: _business_date(r.get("날짜"), r.get("시간")),
            axis=1
        ) if not sorted_detail.empty else []
        round_dates = [
            d for d in sorted(sorted_detail["정산일자"].dropna().unique())
        ] if not sorted_detail.empty else []
        round_map = {d: f"{idx}회차" for idx, d in enumerate(round_dates, start=1)}
        if not sorted_detail.empty:
            sorted_detail["회차"] = sorted_detail["정산일자"].map(round_map).fillna("")
        round_names = [round_map[d] for d in round_dates]

    heart_by_round = {}
    normal_total = 0
    partner_total = 0
    if not sorted_detail.empty:
        heart_by_round = (
            sorted_detail.groupby("회차")["후원하트"]
            .sum()
            .to_dict()
        )
        normal_total = int(sorted_detail.loc[sorted_detail["구분"] == "일반", "후원하트"].sum())
        partner_total = int(sorted_detail.loc[sorted_detail["구분"] == "제휴", "후원하트"].sum())

    cached_values = {"J4": 45}

    headers = [" ", "수량", "정산금", "상/벌금", "헤메", "총 정산금", "비고"]
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=value)
        style_header(cell)
    ws["C3"] = '=TEXT($J$3,"0%")&" 정산금"'

    first_round_row = 4
    round_count = max(len(round_names), 1)
    for offset in range(round_count):
        row = first_round_row + offset
        round_name = round_names[offset] if round_names else "1회차"
        round_heart = int(heart_by_round.get(round_name, 0))
        round_amount = int(round_heart * 45)
        ws.cell(row=row, column=1, value=round_name)
        ws.cell(row=row, column=2, value=f'=SUMIF(\'후원내역\'!A:A,\'정산시트\'!A{row},\'후원내역\'!F:F)')
        ws.cell(row=row, column=3, value=f"=B{row}*$J$3*100")
        ws.cell(row=row, column=6, value=f"=C{row}+D{row}+E{row}")
        cached_values[f"B{row}"] = round_heart
        cached_values[f"C{row}"] = round_amount
        cached_values[f"F{row}"] = round_amount
        ws.cell(row=row, column=7, value="")
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.font = normal_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in (2, 3, 4, 5, 6):
            ws.cell(row=row, column=col).number_format = "#,##0"

    total_row = first_round_row + round_count
    ws.cell(row=total_row, column=1, value="합계")
    ws.cell(row=total_row, column=2, value=f"=SUM(B{first_round_row}:B{total_row - 1})")
    ws.cell(row=total_row, column=3, value=f"=SUM(C{first_round_row}:C{total_row - 1})")
    ws.cell(row=total_row, column=4, value=f"=SUM(D{first_round_row}:D{total_row - 1})")
    ws.cell(row=total_row, column=5, value=f"=SUM(E{first_round_row}:E{total_row - 1})")
    ws.cell(row=total_row, column=6, value=f"=SUM(F{first_round_row}:F{total_row - 1})")
    total_heart = int(sum(heart_by_round.get(round_name, 0) for round_name in round_names))
    total_amount = int(total_heart * 45)
    cached_values[f"B{total_row}"] = total_heart
    cached_values[f"C{total_row}"] = total_amount
    cached_values[f"D{total_row}"] = 0
    cached_values[f"E{total_row}"] = 0
    cached_values[f"F{total_row}"] = total_amount
    for col in range(1, 8):
        cell = ws.cell(row=total_row, column=col)
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in (2, 3, 4, 5, 6):
        ws.cell(row=total_row, column=col).number_format = "#,##0"

    summary_header_row = total_row + 3
    summary_headers = ["일자", "구분", "하트 개수", "공급가액", "세액", "합계", "비고"]
    for col, value in enumerate(summary_headers, start=1):
        cell = ws.cell(row=summary_header_row, column=col, value=value)
        style_header(cell)

    rows = [
        ("일반하트", '=SUMIF(\'후원내역\'!G:G,"일반",\'후원내역\'!F:F)', "=C{row}*$J$3*100", "", ""),
        ("협력지원금", "=C{normal_row}", "=C{row}*IF($J$5>1,$J$5/100,$J$5)*100", "", "J5 협력지원율 기준"),
        ("제휴하트", '=SUMIF(\'후원내역\'!G:G,"제휴",\'후원내역\'!F:F)', "=C{row}*$J$3*100", "", ""),
        ("헤메", "", "", "=D{row}*0.1", ""),
        ("상/벌금", "", "", "=D{row}*0.1", "상벌금 합계"),
    ]
    normal_heart_row = summary_header_row + 1

    for idx, (label, heart_formula, supply_formula, tax_formula, note) in enumerate(rows, start=1):
        row = summary_header_row + idx
        ws.cell(row=row, column=2, value=label)
        ws.cell(row=row, column=2).fill = yellow_fill
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
        if heart_formula:
            ws.cell(row=row, column=3, value=heart_formula.format(normal_row=normal_heart_row, row=row))
        if supply_formula:
            ws.cell(row=row, column=4, value=supply_formula.format(normal_row=normal_heart_row, row=row))
        if tax_formula:
            ws.cell(row=row, column=5, value=tax_formula.format(row=row))
        else:
            ws.cell(row=row, column=5, value=f"=D{row}*0.1")
        ws.cell(row=row, column=6, value=f"=D{row}+E{row}")
        ws.cell(row=row, column=7, value=note)
        for col in range(3, 7):
            ws.cell(row=row, column=col).number_format = "#,##0"
        if label == "협력지원금":
            ws.cell(row=row, column=3).number_format = "#,##0"
        if label == "일반하트":
            amount = int(normal_total * 45)
            tax = int(amount * 0.1)
            cached_values[f"C{row}"] = normal_total
            cached_values[f"D{row}"] = amount
            cached_values[f"E{row}"] = tax
            cached_values[f"F{row}"] = amount + tax
        elif label == "협력지원금":
            amount = int(normal_total * 5)
            tax = int(amount * 0.1)
            cached_values[f"C{row}"] = normal_total
            cached_values[f"D{row}"] = amount
            cached_values[f"E{row}"] = tax
            cached_values[f"F{row}"] = amount + tax
        elif label == "제휴하트":
            amount = int(partner_total * 45)
            tax = int(amount * 0.1)
            cached_values[f"C{row}"] = partner_total
            cached_values[f"D{row}"] = amount
            cached_values[f"E{row}"] = tax
            cached_values[f"F{row}"] = amount + tax
        else:
            cached_values[f"E{row}"] = 0
            cached_values[f"F{row}"] = 0

    final_row = summary_header_row + len(rows) + 1
    ws.cell(row=final_row, column=1, value="합계")
    ws.cell(row=final_row, column=3, value=f"=C{summary_header_row + 1}+C{summary_header_row + 3}")
    ws.cell(row=final_row, column=4, value=f"=SUM(D{summary_header_row + 1}:D{final_row - 1})")
    ws.cell(row=final_row, column=5, value=f"=SUM(E{summary_header_row + 1}:E{final_row - 1})")
    ws.cell(row=final_row, column=6, value=f"=SUM(F{summary_header_row + 1}:F{final_row - 1})")
    support_amount = int(normal_total * 5)
    final_heart = normal_total + partner_total
    final_supply = int((normal_total * 45) + support_amount + (partner_total * 45))
    final_tax = int(final_supply * 0.1)
    cached_values[f"C{final_row}"] = final_heart
    cached_values[f"D{final_row}"] = final_supply
    cached_values[f"E{final_row}"] = final_tax
    cached_values[f"F{final_row}"] = final_supply + final_tax
    for col in range(1, 8):
        cell = ws.cell(row=final_row, column=col)
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(3, 7):
        ws.cell(row=final_row, column=col).number_format = "#,##0"

    for col, width in {
        "A": 15, "B": 14, "C": 16, "D": 16, "E": 14, "F": 16, "G": 30,
        "I": 14, "J": 12,
    }.items():
        ws.column_dimensions[col].width = width
    for row in range(1, final_row + 1):
        ws.row_dimensions[row].height = 22
    apply_border(ws)

    log_ws.append(["회차", "날짜", "시간", "아이디", "닉네임", "하트", "구분"])
    for col in range(1, 8):
        style_header(log_ws.cell(row=1, column=col))

    for _, r in sorted_detail.iterrows():
        row = log_ws.max_row + 1
        log_ws.cell(row=row, column=1, value=r.get("회차", ""))
        log_ws.cell(row=row, column=2, value=r.get("날짜"))
        log_ws.cell(row=row, column=3, value=r.get("시간"))
        log_ws.cell(row=row, column=4, value=r.get("아이디"))
        log_ws.cell(row=row, column=5, value=r.get("닉네임"))
        heart = pd.to_numeric(r.get("후원하트", 0), errors="coerce")
        heart = 0 if pd.isna(heart) else int(max(heart, 0))
        log_ws.cell(row=row, column=6, value=heart)
        log_ws.cell(row=row, column=7, value=r.get("구분"))
        log_ws.cell(row=row, column=6).number_format = "#,##0"

    log_ws["B1"] = "날짜"
    log_ws.freeze_panes = "A2"
    for col, width in {
        "A": 12, "B": 14, "C": 12, "D": 28, "E": 24, "F": 14, "G": 12,
    }.items():
        log_ws.column_dimensions[col].width = width
    apply_border(log_ws)

    return _save_workbook_with_cached_values(wb, cached_values)


def safe_filename(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]+', "_", str(name)).strip() or "download"


def make_downloads_zip(files: list[tuple[str, BytesIO]]) -> BytesIO:
    bio = BytesIO()
    with zipfile.ZipFile(bio, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for filename, file_data in files:
            file_data.seek(0)
            zf.writestr(filename, file_data.read())
            file_data.seek(0)
    bio.seek(0)
    return bio


# ==================================================
# 📥 다운로드 UI
# ==================================================
st.success("집계 완료")

settlement_files = []
bj_files = []
standard_settlement_files = []

# 여러 파일 업로드일 때만 총합산 제공(요구사항)
if len(uploaded_files) > 1:
    total_file = make_total_excel(merged)
    if total_file is None:
        st.warning("총합산 생성 실패: 필수 컬럼(후원시간/후원아이디/후원하트/참여BJ)을 찾지 못했습니다.")
    else:
        st.download_button(
            label="총합산.xlsx 다운로드",
            data=total_file,
            file_name="총합산.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

for bj, views in result.items():
    safe_bj = safe_filename(bj)

    filename1 = (
        f"{prefix}_{safe_bj}_정산용.xlsx"
        if prefix else
        f"{safe_bj}_정산용.xlsx"
    )

    filename2 = (
        f"{prefix}_{safe_bj}_BJ용.xlsx"
        if prefix else
        f"{safe_bj}_BJ용.xlsx"
    )

    filename3 = (
        f"{prefix}_{safe_bj}_표준정산시트.xlsx"
        if prefix else
        f"{safe_bj}_표준정산시트.xlsx"
    )

    settlement_files.append((
        filename1,
        make_excel(
            views["정산용"],
            bj,
            views.get("전체로그")
        )
    ))

    standard_settlement_files.append((
        filename3,
        make_standard_settlement_excel(
            views.get("전체로그"),
            bj,
            round_labels if len(uploaded_files) > 1 else None
        )
    ))

    bj_files.append((
        filename2,
        make_excel(
            views["BJ용"],
            bj,
            views.get("전체로그")
        )
    ))

if settlement_files:
    zip_name = f"{prefix}_정산용_전체다운로드.zip" if prefix else "정산용_전체다운로드.zip"
    st.download_button(
        label="정산용 전체 ZIP 다운로드",
        data=make_downloads_zip(settlement_files),
        file_name=zip_name,
        mime="application/zip"
    )

if bj_files:
    zip_name = f"{prefix}_BJ용_전체다운로드.zip" if prefix else "BJ용_전체다운로드.zip"
    st.download_button(
        label="BJ용 전체 ZIP 다운로드",
        data=make_downloads_zip(bj_files),
        file_name=zip_name,
        mime="application/zip"
    )

if standard_settlement_files:
    zip_name = f"{prefix}_표준정산시트_전체다운로드.zip" if prefix else "표준정산시트_전체다운로드.zip"
    st.download_button(
        label="표준정산시트 전체 ZIP 다운로드",
        data=make_downloads_zip(standard_settlement_files),
        file_name=zip_name,
        mime="application/zip"
    )

# BJ별 파일 제공 (파일 1개일 때만 prefix 붙임)
for bj, views in result.items():

    st.subheader(bj)

    filename1 = (
        f"{prefix}_{safe_filename(bj)}_정산용.xlsx"
        if prefix else
        f"{safe_filename(bj)}_정산용.xlsx"
    )

    filename2 = (
        f"{prefix}_{safe_filename(bj)}_BJ용.xlsx"
        if prefix else
        f"{safe_filename(bj)}_BJ용.xlsx"
    )

    filename3 = (
        f"{prefix}_{safe_filename(bj)}_표준정산시트.xlsx"
        if prefix else
        f"{safe_filename(bj)}_표준정산시트.xlsx"
    )

    file1 = next(data for name, data in settlement_files if name == filename1)
    file2 = next(data for name, data in bj_files if name == filename2)
    file3 = next(data for name, data in standard_settlement_files if name == filename3)

    st.download_button(
        label=f"{filename1} 다운로드",
        data=file1,
        file_name=filename1,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.download_button(
        label=f"{filename2} 다운로드",
        data=file2,
        file_name=filename2,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.download_button(
        label=f"{filename3} 다운로드",
        data=file3,
        file_name=filename3,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
